import io
import base64
import datetime
from odoo import models, fields, api
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class BinCardItemWiseWizard(models.TransientModel):
    _name = 'bin.card.item.wise.wizard'
    _description = 'BIN Card Item Wise Report Wizard'

    date_from    = fields.Date(string='Date From', required=True, default=fields.Date.context_today)
    date_to      = fields.Date(string='Date To',   required=True, default=fields.Date.context_today)
    product_ids  = fields.Many2many('product.product', string='Products',
                                    help='Leave empty to include all products')
    warehouse_id = fields.Many2one('stock.warehouse', string='Warehouse')
    location_id  = fields.Many2one('stock.location', string='Location')

    # Computed domain field — drives the location_id domain in the view
    location_domain = fields.Binary(
        string='Location Domain',
        compute='_compute_location_domain',
    )

    @api.depends('warehouse_id')
    def _compute_location_domain(self):
        for rec in self:
            if rec.warehouse_id:
                child_locs = self.env['stock.location'].search([
                    ('id', 'child_of', rec.warehouse_id.view_location_id.id),
                    ('usage', '=', 'internal'),
                ])
                rec.location_domain = [('id', 'in', child_locs.ids)]
            else:
                rec.location_domain = [('usage', '=', 'internal')]

    @api.onchange('warehouse_id')
    def _onchange_warehouse_id(self):
        self.location_id = False

    # ── helpers ───────────────────────────────────────────────────────────────
    def _fill(self, hex_color):
        return PatternFill(fill_type='solid', fgColor=hex_color)

    def _style_header(self, cell, bg='1F3864'):
        s = Side(style='thin')
        cell.font      = Font(name='Arial', size=9, bold=True, color='FFFFFF')
        cell.fill      = self._fill(bg)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border    = Border(left=s, right=s, top=s, bottom=s)

    def _style_data(self, cell, horizontal='center', bold=False,
                    bg=None, number_format=None, font_color='000000'):
        s = Side(style='thin')
        cell.font      = Font(name='Arial', size=9, bold=bold, color=font_color)
        cell.alignment = Alignment(horizontal=horizontal, vertical='center')
        cell.border    = Border(left=s, right=s, top=s, bottom=s)
        if bg:
            cell.fill = self._fill(bg)
        if number_format:
            cell.number_format = number_format

    def _get_trans_type(self, move):
        if move.picking_type_id:
            return {'incoming': 'GRN', 'outgoing': 'INV', 'internal': 'INT'}.get(
                move.picking_type_id.code, move.picking_type_id.name)
        return 'ADJ'

    # ── main action ───────────────────────────────────────────────────────────
    def action_print_report(self):
        self.ensure_one()
        wb = Workbook()
        ws = wb.active
        ws.title = 'BIN CARD Item Wise'

        num_fmt  = '#,##0.00'
        date_fmt = 'DD/MM/YYYY'

        for i, w in enumerate([14, 14, 26, 14, 26, 12, 12], start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

        # Row 1 — Title
        ws.row_dimensions[1].height = 28
        ws.merge_cells('A1:G1')
        ws['A1'].value     = 'BIN CARD - ITEM WISE REPORT'
        ws['A1'].font      = Font(name='Arial', size=14, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

        # Row 2 — Meta
        ws.row_dimensions[2].height = 18
        ws['A2'].value = 'Company:'
        ws['A2'].font  = Font(name='Arial', size=9, bold=True)
        ws.merge_cells('B2:C2')
        display_company = ", ".join(report_companies.mapped('name'))
        ws['B2'].value = display_company
        self._style_data(ws['B2'], horizontal='left')

        ws['D2'].value = 'Date From:'
        ws['D2'].font  = Font(name='Arial', size=9, bold=True)
        ws['E2'].value = self.date_from
        self._style_data(ws['E2'], number_format=date_fmt)

        ws['F2'].value = 'Date To:'
        ws['F2'].font  = Font(name='Arial', size=9, bold=True)
        ws['G2'].value = self.date_to
        self._style_data(ws['G2'], number_format=date_fmt)

        # Row 4 — Column headers
        ws.row_dimensions[4].height = 22
        headers = ['Date', 'Item Code', 'Item Name', 'Transactions', 'Reference No', 'QTY', 'Balance']
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=4, column=col_idx, value=header)
            self._style_header(cell, '1F3864')

        # Fetch Data
        report_companies = self.env.companies
        if self.location_id and self.location_id.company_id:
            report_companies = self.location_id.company_id
        elif self.warehouse_id and self.warehouse_id.company_id:
            report_companies = self.warehouse_id.company_id

        company_ids = report_companies.ids
        if self.location_id:
            target_loc_ids = [self.location_id.id]
        elif self.warehouse_id:
            target_loc_ids = self.env['stock.location'].search([
                ('id', 'child_of', self.warehouse_id.view_location_id.id),
                ('usage', '=', 'internal'),
                ('company_id', 'in', company_ids)
            ]).ids
        else:
            target_loc_ids = self.env['stock.location'].search([
                ('usage', '=', 'internal'),
                ('company_id', 'in', company_ids)
            ]).ids

        start_dt = str(self.date_from) + ' 00:00:00'
        end_dt   = str(self.date_to)   + ' 23:59:59'
        
        # Build move domain
        move_domain = [
            ('state', '=', 'done'),
            ('date', '<=', end_dt),
            '|',
            ('location_id', 'in', target_loc_ids),
            ('location_dest_id', 'in', target_loc_ids),
        ]
        # 1. Build Move Domain (Location & Date are always filtered)
        move_domain = [
            ('state', '=', 'done'),
            ('date', '<=', end_dt),
            '|',
            ('location_id', 'in', target_loc_ids),
            ('location_dest_id', 'in', target_loc_ids),
        ]
        if self.product_ids:
            move_domain.append(('product_id', 'in', self.product_ids.ids))
        
        # 2. Fetch all moves in bulk
        move_fields = ['product_id', 'product_qty', 'date', 'location_id', 'location_dest_id', 'reference', 'origin', 'picking_type_id']
        all_moves = self.env['stock.move'].search_read(move_domain, move_fields)

        # 3. Determine products to process
        if self.product_ids:
            products = self.product_ids.sorted(key=lambda p: (p.default_code or '', p.name or ''))
        else:
            # If no products selected, process all products that have moves in these locations
            # This covers both Opening Balance (moves before start_dt) and Period Transactions.
            move_product_ids = list({m['product_id'][0] for m in all_moves})
            products = self.env['product.product'].browse(move_product_ids).sorted(
                key=lambda p: (p.default_code or '', p.name or '')
            )

        # 4. Group moves by product
        moves_by_product = {}
        for m in all_moves:
            p_id = m['product_id'][0]
            if p_id not in moves_by_product:
                moves_by_product[p_id] = []
            moves_by_product[p_id].append(m)

        data_row = 5
        for product in products:
            p_moves = moves_by_product.get(product.id, [])
            # Sort by date for correct running balance
            p_moves.sort(key=lambda x: x['date'])

            running_bal = 0.0
            period_moves = []
            
            for m in p_moves:
                m_date = str(m['date'])
                if m_date < start_dt:
                    if m['location_dest_id'][0] in target_loc_ids: running_bal += m['product_qty']
                    if m['location_id'][0]      in target_loc_ids: running_bal -= m['product_qty']
                else:
                    period_moves.append(m)

            # Skip products with no moves and no balance
            if not period_moves and running_bal == 0:
                continue

            # 1. Row for Opening Balance (Optional but helpful if we want to show non-zero opening balances)
            # Actually, the user asked to "load all product", but typically in item-wise 
            # we only show transactions. I will add an opening balance row if it's non-zero 
            # OR if there are period moves.
            
            # Show Opening Balance Row if non-zero
            if running_bal != 0 or period_moves:
                row_values = [
                    self.date_from,
                    product.default_code or '',
                    product.name,
                    'Opening Balance',
                    '',
                    '',
                    running_bal,
                ]
                ws.row_dimensions[data_row].height = 18
                for col_idx, val in enumerate(row_values, start=1):
                    cell = ws.cell(row=data_row, column=col_idx, value=val)
                    if col_idx == 1:
                        self._style_data(cell, number_format=date_fmt)
                    elif col_idx in (2, 3, 5):
                        self._style_data(cell, horizontal='left')
                    elif col_idx in (6, 7):
                        self._style_data(cell, horizontal='right', number_format=num_fmt)
                    else:
                        self._style_data(cell, bold=True)
                data_row += 1

            # 2. Rows for Period Moves
            for move in period_moves:
                is_in  = move['location_dest_id'][0] in target_loc_ids
                is_out = move['location_id'][0]      in target_loc_ids

                if is_in and is_out:
                    qty_change = 0.0
                elif is_in:
                    qty_change = move['product_qty']
                elif is_out:
                    qty_change = -move['product_qty']
                else:
                    continue

                running_bal += qty_change
                
                # Fetching the move record to use the existing _get_trans_type or similar
                # Since search_read returns a dict, we can either browse or adapt.
                # I'll adapt to use browse once per move to ensure accuracy if needed, 
                # but for performance I'll use the dict data for trans_type if possible.
                # Actually, I'll browse the move for trans_type to keep it "like previous"
                m_rec = self.env['stock.move'].browse(move['id'])
                trans_type = self._get_trans_type(m_rec)
                ref = m_rec.reference or m_rec.origin or ''

                m_dt = move['date']
                if isinstance(m_dt, str):
                    m_dt = datetime.datetime.strptime(m_dt, '%Y-%m-%d %H:%M:%S')

                row_values = [
                    m_dt.date() if hasattr(m_dt, 'date') else m_dt,
                    product.default_code or '',
                    product.name,
                    trans_type,
                    ref,
                    abs(qty_change) if not (is_in and is_out) else 0.0,
                    running_bal,
                ]

                ws.row_dimensions[data_row].height = 18
                for col_idx, val in enumerate(row_values, start=1):
                    cell = ws.cell(row=data_row, column=col_idx, value=val)
                    if col_idx == 1:
                        self._style_data(cell, number_format=date_fmt)
                    elif col_idx in (2, 3, 5):
                        self._style_data(cell, horizontal='left')
                    elif col_idx in (6, 7):
                        self._style_data(cell, horizontal='right', number_format=num_fmt)
                    else:
                        self._style_data(cell)
                data_row += 1

        # Save & return
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        file_data = base64.b64encode(output.read())
        output.close()

        att = self.env['ir.attachment'].create({
            'name':     f'BIN_Card_ItemWise_{self.date_from}_{self.date_to}.xlsx',
            'type':     'binary',
            'datas':    file_data,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        })
        return {
            'type':   'ir.actions.act_url',
            'url':    f'/web/content/{att.id}?download=true',
            'target': 'self',
        }