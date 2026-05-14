import io
import base64
from odoo import models, fields, api
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class BinCardWizard(models.TransientModel):
    _name = 'bin.card.wizard'
    _description = 'BIN Card Report Wizard'

    date_from    = fields.Date(string='Date From', required=True, default=fields.Date.context_today)
    date_to      = fields.Date(string='Date To',   required=True, default=fields.Date.context_today)
    warehouse_id = fields.Many2one('stock.warehouse', string='Warehouse')
    location_id  = fields.Many2one('stock.location', string='Location')

    # Computed domain field — drives the location_id domain in the view
    location_domain = fields.Binary(
        string='Location Domain',
        compute='_compute_location_domain',
    )

    @api.depends('warehouse_id')
    def _compute_location_domain(self):
        for rec in self:
            if rec.warehouse_id:
                child_locs = self.env['stock.location'].search([
                    ('id', 'child_of', rec.warehouse_id.view_location_id.id),
                    ('usage', '=', 'internal'),
                ])
                rec.location_domain = [('id', 'in', child_locs.ids)]
            else:
                rec.location_domain = [('usage', '=', 'internal')]

    @api.onchange('warehouse_id')
    def _onchange_warehouse_id(self):
        self.location_id = False

    # ── helpers ───────────────────────────────────────────────────────────────
    def _thin_border(self):
        s = Side(style='thin')
        return Border(left=s, right=s, top=s, bottom=s)

    def _fill(self, hex_color):
        return PatternFill(fill_type='solid', fgColor=hex_color)

    def _apply(self, cell, font=None, fill=None, align=None, border=None, number_format=None):
        if font:          cell.font          = font
        if fill:          cell.fill          = fill
        if align:         cell.alignment     = align
        if border:        cell.border        = border
        if number_format: cell.number_format = number_format

    def _style_header(self, cell, bg='1F3864'):
        self._apply(
            cell,
            font=Font(name='Arial', size=9, bold=True, color='FFFFFF'),
            fill=self._fill(bg),
            align=Alignment(horizontal='center', vertical='center', wrap_text=True),
            border=self._thin_border(),
        )

    def _style_data(self, cell, horizontal='center', number_format=None):
        self._apply(
            cell,
            font=Font(name='Arial', size=9),
            align=Alignment(horizontal=horizontal, vertical='center'),
            border=self._thin_border(),
            number_format=number_format,
        )

    def _get_trans_type(self, move):
        if move.picking_type_id:
            return {'incoming': 'GRN', 'outgoing': 'INV', 'internal': 'INT'}.get(
                move.picking_type_id.code, move.picking_type_id.name)
        return 'ADJ'

    def _get_opening_balance(self, product_id, move_date, location_id):
        prior = self.env['stock.move'].search([
            ('product_id', '=', product_id),
            ('state', '=', 'done'),
            ('date', '<', move_date),
        ])
        qty = 0.0
        for m in prior:
            if location_id:
                if m.location_dest_id == location_id:   qty += m.product_qty
                elif m.location_id    == location_id:   qty -= m.product_qty
            else:
                if m.location_dest_id.usage == 'internal': qty += m.product_qty
                if m.location_id.usage      == 'internal': qty -= m.product_qty
        return qty

    # ── main action ───────────────────────────────────────────────────────────
    def action_print_report(self):
        self.ensure_one()
        wb = Workbook()
        ws = wb.active
        ws.title = 'BIN CARD'

        dark_blue   = '1F3864'
        mid_blue    = '2E75B6'
        center      = Alignment(horizontal='center', vertical='center')
        num_fmt     = '#,##0.00'

        # ── Data Collection & Analysis ────────────────────────────────────────
        report_companies = self.env.companies
        if self.location_id and self.location_id.company_id:
            report_companies = self.location_id.company_id
        elif self.warehouse_id and self.warehouse_id.company_id:
            report_companies = self.warehouse_id.company_id
            
        company_ids = report_companies.ids
        
        if self.location_id:
            target_loc_ids = [self.location_id.id]
        elif self.warehouse_id:
            target_loc_ids = self.env['stock.location'].search([
                ('id', 'child_of', self.warehouse_id.view_location_id.id),
                ('usage', '=', 'internal'),
                ('company_id', 'in', company_ids)
            ]).ids
        else:
            target_loc_ids = self.env['stock.location'].search([
                ('usage', '=', 'internal'),
                ('company_id', 'in', company_ids)
            ]).ids

        start_dt = str(self.date_from) + ' 00:00:00'
        end_dt   = str(self.date_to)   + ' 23:59:59'

        move_domain = [
            ('state', '=', 'done'),
            ('date', '<=', end_dt),
            '|',
            ('location_id', 'in', target_loc_ids),
            ('location_dest_id', 'in', target_loc_ids),
        ]
        move_fields = ['product_id', 'product_qty', 'date', 'location_id', 'location_dest_id']
        all_moves = self.env['stock.move'].search_read(move_domain, move_fields)

        moves_by_product = {}
        for m in all_moves:
            p_id = m['product_id'][0]
            if p_id not in moves_by_product:
                moves_by_product[p_id] = []
            moves_by_product[p_id].append(m)

        product_ids_to_process = list(moves_by_product.keys())
        products = self.env['product.product'].browse(product_ids_to_process).sorted(
            key=lambda p: (p.default_code or '', p.name or '')
        )

        loc_ids = set()
        for m in all_moves:
            loc_ids.add(m['location_id'][0])
            loc_ids.add(m['location_dest_id'][0])
        loc_data = {l.id: l.usage for l in self.env['stock.location'].browse(list(loc_ids))}

        # Pre-calculate data per product.
        # A product row is included only if it has at least one movement
        # (receipt, issue, or adjustment) within the selected date range.
        # Receipts, Issues, Adjustments columns are ALWAYS shown.
        all_rows = []

        for product in products:
            p_moves = moves_by_product.get(product.id, [])
            open_qty, qty_receipt, qty_issue, qty_adj = 0.0, 0.0, 0.0, 0.0

            for m in p_moves:
                m_date = str(m['date'])
                is_in  = m['location_dest_id'][0] in target_loc_ids
                is_out = m['location_id'][0]      in target_loc_ids

                if m_date < start_dt:
                    if is_in:  open_qty += m['product_qty']
                    if is_out: open_qty -= m['product_qty']
                else:
                    if is_in and is_out: continue
                    if is_in:
                        src_usage = loc_data.get(m['location_id'][0])
                        if src_usage == 'supplier': qty_receipt += m['product_qty']
                        elif src_usage in ('inventory', 'production'): qty_adj += m['product_qty']
                        else: qty_receipt += m['product_qty']
                    elif is_out:
                        dest_usage = loc_data.get(m['location_dest_id'][0])
                        if dest_usage == 'customer': qty_issue += m['product_qty']
                        elif dest_usage in ('inventory', 'production'): qty_adj -= m['product_qty']
                        else: qty_issue += m['product_qty']

            closing_qty = open_qty + qty_receipt - qty_issue + qty_adj

            # Skip rows with no activity at all
            if open_qty == 0 and qty_receipt == 0 and qty_issue == 0 and qty_adj == 0 and closing_qty == 0:
                continue

            # Skip rows with no period movement (zero Receipts, Issues AND Adjustments)
            # — only opening/closing balance present means no actual transactions to show
            if qty_receipt == 0 and qty_issue == 0 and qty_adj == 0:
                continue

            all_rows.append({
                'product': product,
                'open': open_qty,
                'receipt': qty_receipt,
                'issue': qty_issue,
                'adj': qty_adj,
                'close': closing_qty,
                'price': product.standard_price,
            })

        # ── Column definitions — always show all movement columns ──────────────
        col_groups = [
            ('Open Balance',    'open'),
            ('Receipts',        'receipt'),
            ('Issues',          'issue'),
            ('Adjustments',     'adj'),
            ('Closing Balance', 'close'),
        ]

        total_cols = 3 + (len(col_groups) * 3)
        last_col_letter = get_column_letter(total_cols)

        # ── Setup Sheet ───────────────────────────────────────────────────────
        col_widths = [14, 24, 10] + ([12, 12, 12] * len(col_groups))
        for i, w in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

        ws.merge_cells(f'A1:{last_col_letter}1')
        title_cell = ws['A1']
        title_cell.value = 'BIN CARD REPORT'
        title_cell.font = Font(name='Arial', size=14, bold=True)
        title_cell.alignment = center

        # Row 2 (Meta) stays mostly same
        ws['A2'].value = 'Company:'; ws['A2'].font = Font(name='Arial', size=9, bold=True)
        ws.merge_cells('B2:C2')
        ws['B2'].value = ", ".join(report_companies.mapped('name'))
        self._style_data(ws['B2'], horizontal='left')
        ws['D2'].value = 'Date From:'; ws['D2'].font = Font(name='Arial', size=9, bold=True)
        ws['E2'].value = self.date_from; self._style_data(ws['E2'], number_format='DD/MM/YYYY')
        ws['F2'].value = 'Date To:'; ws['F2'].font = Font(name='Arial', size=9, bold=True)
        ws['G2'].value = self.date_to; self._style_data(ws['G2'], number_format='DD/MM/YYYY')

        # Row 4 & 5 (Headers)
        ws.row_dimensions[4].height = 32
        for col_idx, label in enumerate(['Item Code', 'Item Name', 'UOM'], start=1):
            ws.merge_cells(start_row=4, start_column=col_idx, end_row=5, end_column=col_idx)
            self._style_header(ws.cell(row=4, column=col_idx, value=label), dark_blue)

        current_col = 4
        for label, _key in col_groups:
            start_col = get_column_letter(current_col)
            end_col   = get_column_letter(current_col + 2)
            ws.merge_cells(f'{start_col}4:{end_col}4')
            ws[f'{start_col}4'].value = label
            self._style_header(ws[f'{start_col}4'], mid_blue)
            
            for sub_idx, sub_label in enumerate(['Quantity', 'Per Value', 'Value']):
                cell = ws.cell(row=5, column=current_col + sub_idx, value=sub_label)
                self._style_header(cell, mid_blue)
            current_col += 3

        # ── Populate Data Rows ────────────────────────────────────────────────
        data_row_idx = 6
        for row in all_rows:
            p = row['product']
            row_data = [p.default_code or '', p.name, p.uom_id.name]
            
            for _label, key in col_groups:
                qty = row[key]
                val = row['price']
                row_data.extend([qty, val, qty * val])

            ws.row_dimensions[data_row_idx].height = 18
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=data_row_idx, column=col_idx, value=value)
                if col_idx <= 3:
                    self._style_data(cell, horizontal='left' if col_idx == 2 else 'center')
                else:
                    self._style_data(cell, horizontal='right', number_format=num_fmt)
            data_row_idx += 1

        # ── Save & return ─────────────────────────────────────────────────────
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        file_data = base64.b64encode(output.read())
        output.close()

        att = self.env['ir.attachment'].create({
            'name':     f'BIN_Card_{self.date_from}_{self.date_to}.xlsx',
            'type':     'binary',
            'datas':    file_data,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        })
        return {
            'type':   'ir.actions.act_url',
            'url':    f'/web/content/{att.id}?download=true',
            'target': 'self',
        }