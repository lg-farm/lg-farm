import io
import base64
from odoo import models, fields, api
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class DaySalesDetailsWizard(models.TransientModel):
    _name = 'day.sales.details.wizard'
    _description = 'Day Sales Details Report Wizard'

    date_from = fields.Date(string='Date From', required=True, default=fields.Date.context_today)
    date_to = fields.Date(string='Date To', required=True, default=fields.Date.context_today)
    company_ids = fields.Many2many('res.company', string='Companies', 
                                   domain=lambda self: [('id', 'in', self.env.companies.ids)])

    def _thin_border(self):
        s = Side(style='thin')
        return Border(left=s, right=s, top=s, bottom=s)

    def _fill(self, hex_color):
        return PatternFill(fill_type='solid', fgColor=hex_color)

    def _apply(self, cell, font=None, fill=None, align=None, border=None, number_format=None):
        if font:          cell.font          = font
        if fill:          cell.fill          = fill
        if align:         cell.alignment     = align
        if border:        cell.border        = border
        if number_format: cell.number_format = number_format

    def _style_header(self, cell, bg='1F3864'):
        self._apply(
            cell,
            font=Font(name='Arial', size=9, bold=True, color='FFFFFF'),
            fill=self._fill(bg),
            align=Alignment(horizontal='center', vertical='center', wrap_text=True),
            border=self._thin_border(),
        )

    def _style_data(self, cell, horizontal='center', number_format=None, bold=False):
        self._apply(
            cell,
            font=Font(name='Arial', size=9, bold=bold),
            align=Alignment(horizontal=horizontal, vertical='center'),
            border=self._thin_border(),
            number_format=number_format,
        )

    def action_generate_excel(self):
        self.ensure_one()
        wb = Workbook()
        ws = wb.active
        ws.title = 'Day Sales Details'

        border      = self._thin_border()
        dark_blue   = '1F3864'
        mid_blue    = '2E75B6'
        center      = Alignment(horizontal='center', vertical='center')
        left        = Alignment(horizontal='left',   vertical='center')
        right       = Alignment(horizontal='right',  vertical='center')
        num_fmt     = '#,##0.00'

        # Column widths
        col_widths = [25, 15, 12, 12, 12, 12, 15, 15, 15, 15, 15]
        for i, w in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

        # Title
        ws.merge_cells('A1:K1')
        title_cell = ws['A1']
        title_cell.value = 'DAY SALES DETAILS'
        title_cell.font = Font(name='Arial', size=12, bold=True)
        title_cell.alignment = center

        # Meta info
        ws['A4'] = 'DAY SALES DETAILS'
        ws['A4'].font = Font(bold=True)
        ws['A5'] = 'DATE FROM:'
        ws['B5'] = self.date_from
        self._style_data(ws['B5'], horizontal='left', number_format='DD-MM-YYYY')
        ws['D5'] = 'DATE TO:'
        ws['E5'] = self.date_to
        self._style_data(ws['E5'], horizontal='left', number_format='DD-MM-YYYY')

        # Table Headers
        company_ids = self.company_ids.ids or self.env.companies.ids
        currency_name = (self.company_ids[0].currency_id.name if self.company_ids else self.env.company.currency_id.name) or 'Rs.'
        
        headers_row = 7
        ws.merge_cells('A7:A8')
        ws['A7'] = 'Names of Factory Outlets'
        self._style_header(ws['A7'], mid_blue)

        ws.merge_cells('B7:B8')
        ws['B7'] = 'Opening Balance'
        self._style_header(ws['B7'], mid_blue)

        ws.merge_cells('C7:F7')
        ws['C7'] = f'Sales Types ({currency_name})'
        self._style_header(ws['C7'], mid_blue)
        for col in ['C', 'D', 'E', 'F']:
            self._style_header(ws[f'{col}8'], mid_blue)
        ws['C8'] = 'Credit Sales'
        ws['D8'] = 'Cheque Sales'
        ws['E8'] = 'Card Sales'
        ws['F8'] = 'Cash Sales'

        ws.merge_cells('G7:G8')
        ws['G7'] = 'Total Sales'
        self._style_header(ws['G7'], mid_blue)

        ws.merge_cells('H7:J7')
        ws['H7'] = f'Cash Deposit from Cash Sales ({currency_name})'
        self._style_header(ws['H7'], mid_blue)
        for col in ['H', 'I', 'J']:
            self._style_header(ws[f'{col}8'], mid_blue)
        ws['H8'] = 'Remark'
        ws['I8'] = 'Deposited Date'
        ws['J8'] = 'Deposited Amount'

        ws.merge_cells('K7:K8')
        ws['K7'] = 'Undeposite Amount'
        self._style_header(ws['K7'], mid_blue)

        # Data Fetching
        # Build date boundaries (timezone-naive UTC datetimes for domain filtering)
        from datetime import datetime, date as date_type
        date_from_dt = datetime(self.date_from.year, self.date_from.month, self.date_from.day, 0, 0, 0)
        date_to_dt   = datetime(self.date_to.year,   self.date_to.month,   self.date_to.day,   23, 59, 59)

        pos_configs = self.env['pos.config'].search([('company_id', 'in', company_ids)])

        row = 9
        total_credit = 0.0
        total_cheque = 0.0
        total_card = 0.0
        total_cash = 0.0
        total_sales_all = 0.0
        total_deposited = 0.0
        total_variance = 0.0

        for config in pos_configs:
            # Find sessions for this config on the selected date range
            sessions = self.env['pos.session'].search([
                ('config_id', '=', config.id),
                ('start_at', '>=', fields.Datetime.to_string(date_from_dt)),
                ('start_at', '<=', fields.Datetime.to_string(date_to_dt)),
            ])

            if not sessions:
                # Show a zero row for configs with no sessions in the range
                ws.cell(row=row, column=1, value=config.name)
                for c in range(2, 12):
                    ws.cell(row=row, column=c, value=0.0 if c != 8 and c != 9 else '')
                self._style_data(ws.cell(row=row, column=1), horizontal='left')
                for c in range(2, 8):
                    self._style_data(ws.cell(row=row, column=c), horizontal='right', number_format=num_fmt)
                self._style_data(ws.cell(row=row, column=8), horizontal='left')
                self._style_data(ws.cell(row=row, column=9), horizontal='center', number_format='DD-MM-YYYY')
                self._style_data(ws.cell(row=row, column=10), horizontal='right', number_format=num_fmt)
                self._style_data(ws.cell(row=row, column=11), horizontal='right', number_format=num_fmt)
                row += 1
                continue

            # ---------- One row per session ----------
            for session in sessions:
                opening_balance = session.cash_register_balance_start

                # Aggregate payments by type for this session only
                credit = 0.0
                cheque = 0.0
                card   = 0.0
                cash   = 0.0

                orders = self.env['pos.order'].search([('session_id', '=', session.id)])
                for payment in orders.mapped('payment_ids'):
                    method_name = payment.payment_method_id.name.lower()
                    amount = payment.amount
                    if 'cash' in method_name:
                        cash += amount
                    elif 'card' in method_name or 'visa' in method_name or 'master' in method_name:
                        card += amount
                    elif 'cheque' in method_name or 'check' in method_name:
                        cheque += amount
                    elif 'credit' in method_name:
                        credit += amount
                    else:
                        card += amount

                total_sales = credit + cheque + card + cash

                # Cash In / Out Logic for this session
                deposited_amount = 0.0
                remarks = []
                deposited_date = ''

                st_lines = self.env['account.bank.statement.line'].search([
                    ('pos_session_id', '=', session.id),
                    ('amount', '!=', 0)
                ])
                for line in st_lines:
                    is_money_in_out = False
                    if line.payment_ref:
                        ref = line.payment_ref
                        if '-out-' in ref:
                            remarks.append(ref.split('-out-')[-1].strip())
                            is_money_in_out = True
                        elif '-in-' in ref:
                            remarks.append(ref.split('-in-')[-1].strip())
                            is_money_in_out = True

                    if is_money_in_out and line.amount < 0:
                        deposited_amount += abs(line.amount)
                        deposited_date = line.date

                remark_text = ", ".join(set(remarks)) if remarks else ""
                variance = cash - deposited_amount

                # Build the outlet label: "Furniture Shop(00001)"
                # session.name is typically something like "Furniture Shop/00001"
                # Extract the sequence part after the last '/' if present
                session_seq = session.name.split('/')[-1] if '/' in session.name else session.name
                outlet_label = f"{config.name}({session_seq})"

                # Fill Row
                ws.cell(row=row, column=1, value=outlet_label)
                ws.cell(row=row, column=2, value=opening_balance)
                ws.cell(row=row, column=3, value=credit)
                ws.cell(row=row, column=4, value=cheque)
                ws.cell(row=row, column=5, value=card)
                ws.cell(row=row, column=6, value=cash)
                ws.cell(row=row, column=7, value=total_sales)
                ws.cell(row=row, column=8, value=remark_text)
                ws.cell(row=row, column=9, value=deposited_date)
                ws.cell(row=row, column=10, value=deposited_amount)
                ws.cell(row=row, column=11, value=variance)

                # Styling
                self._style_data(ws.cell(row=row, column=1), horizontal='left')
                for c in range(2, 8):
                    self._style_data(ws.cell(row=row, column=c), horizontal='right', number_format=num_fmt)
                self._style_data(ws.cell(row=row, column=8), horizontal='left')
                self._style_data(ws.cell(row=row, column=9), horizontal='center', number_format='DD-MM-YYYY')
                self._style_data(ws.cell(row=row, column=10), horizontal='right', number_format=num_fmt)
                self._style_data(ws.cell(row=row, column=11), horizontal='right', number_format=num_fmt)

                total_credit      += credit
                total_cheque      += cheque
                total_card        += card
                total_cash        += cash
                total_sales_all   += total_sales
                total_deposited   += deposited_amount
                total_variance    += variance
                row += 1

        # Totals Row
        ws.cell(row=row, column=1, value='Total')
        ws.cell(row=row, column=3, value=total_credit)
        ws.cell(row=row, column=4, value=total_cheque)
        ws.cell(row=row, column=5, value=total_card)
        ws.cell(row=row, column=6, value=total_cash)
        ws.cell(row=row, column=7, value=total_sales_all)
        ws.cell(row=row, column=10, value=total_deposited)
        ws.cell(row=row, column=11, value=total_variance)
        
        self._style_data(ws.cell(row=row, column=1), horizontal='left', bold=True)
        for c in [3, 4, 5, 6, 7, 10, 11]:
            self._style_data(ws.cell(row=row, column=c), horizontal='right', number_format=num_fmt, bold=True)
        
        # Bottom borders for empty cells in total row
        for c in [2, 8, 9]:
            self._style_data(ws.cell(row=row, column=c))

        row += 2
        ws.cell(row=row, column=1, value='Prepared By:')
        row += 1
        ws.cell(row=row, column=1, value='Checked By:')
        row += 1
        ws.cell(row=row, column=1, value='Approved By:')

        # Save and return
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        file_data = base64.b64encode(output.read())
        output.close()

        att = self.env['ir.attachment'].create({
            'name':     f'Day_Sales_Details_{self.date_from}_{self.date_to}.xlsx',
            'type':     'binary',
            'datas':    file_data,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        })
        return {
            'type':   'ir.actions.act_url',
            'url':    f'/web/content/{att.id}?download=true',
            'target': 'self',
        }
